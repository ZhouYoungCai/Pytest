1、读取 Excel 文件
	第三方库

	xlrd
	xlwings
	pandas
	openpyxl

	官方文档： https://openpyxl.readthedocs.io/en/stable/
2、openpyxl 库的安装
	安装：pip install openpyxl
	导入：import openpyxl
3、openpyxl 库的操作
	读取工作簿

	读取工作表

	读取单元格

	import openpyxl

	# 获取工作簿
	book = openpyxl.load_workbook('../data/params.xlsx')

	# 读取工作表
	sheet = book.active

	# 读取单个单元格
	cell_a1 = sheet['A1']
	cell_a3 = sheet.cell(column=1, row=3)  # A3

	# 读取多个连续单元格
	cells = sheet["A1":"C3"]

	# 获取单元格的值
	cell_a1.value
4、工程目录结构
	data 目录：存放 excel 数据文件

	func 目录：存放被测函数文件

	testcase 目录：存放测试用例文件

	# 工程目录结构
	.
	├── data
	│   └── params.excel
	├── func
	│   ├── __init__.py
	│   └── operation.py
	└── testcase
		├── __init__.py
		└── test_add.py
5、测试准备
	被测对象：operation.py

	测试用例：test_add.py

	# operation.py 文件内容
	def my_add(x, y):
		result = x + y
		return result

	# test_add.py 文件内容
	class TestWithEXCEL:
		@pytest.mark.parametrize('x,y,expected', get_excel())
		def test_add(self, x, y, expected):
			assert my_add(int(x), int(y)) == int(expected)
6、测试准备
    测试数据：params.xlsx
7、Pytest 数据驱动结合 Excel 文件
	# 读取Excel文件
	import openpyxl
	import pytest

	def get_excel():
		# 获取工作簿
		book = openpyxl.load_workbook('../data/params.xlsx')

		# 获取活动行（非空白的）
		sheet = book.active

		# 提取数据，格式：[[1, 2, 3], [3, 6, 9], [100, 200, 300]]
		values = []
		for row in sheet:
			line = []
			for cell in row:
				line.append(cell.value)
			values.append(line)
		return values

